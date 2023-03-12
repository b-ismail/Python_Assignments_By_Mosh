import openpyxl as xl

# root_path
wb=xl.load_workbook('S:\\Challenge_JH_Tasks\\12_SkipQ\\Python_Assignments_By_Mosh\\49_spreadsheets_manipulation\\trans.xlsx')
# wb=xl.load_workbook('\\trans.xlsx')
sheet = wb['Sheet1']

cell = sheet['a1']
#  OR
for ind in range(1,sheet.max_row+1):
    cell = sheet.cell(ind, 3) # row and column
    # print(cell.value)
    corrected_value = int(cell.value) * 0.9
    corrected_price_cell = sheet.cell(ind, 4)
    corrected_price_cell = corrected_value

wb.save('transactions2.xlsx')